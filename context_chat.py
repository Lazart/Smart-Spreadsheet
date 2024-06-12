import os
import json
import pandas as  pd
from openpyxl import load_workbook
from datetime import date, datetime
from openai import OpenAI
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
# Ensure you've set the OPENAI_API_KEY environment variable in your operating system

api_key = os.getenv('OPENAI_API_KEY')
client = OpenAI(
    api_key=api_key,
)

# Set up CORS middleware
origins = [
    "http://localhost:3000",  # Allow the origin where your frontend runs
    "http://127.0.0.1:3000"   # You can also add this if you access your frontend using 127.0.0.1
]

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)

def has_right_border(cell):
    return cell.border.right.style is not None

def has_left_border(cell):
    return cell.border.left.style is not None

def has_bottom_border(cell):
    return cell.border.bottom.style is not None

def is_blue_header(cell):
    
    fill_color = cell.fill.start_color.index 
    is_blue_background = fill_color == 4  
    return  is_blue_background 

def is_top_left(cell):
    top_left = cell.border.left.style is not None and  cell.border.left.style is not None

    return top_left


def find_last_border(ws, col, max_row, row, top_right, bottom_left, bottom_right):
    for bottom_row in range(row + 1, max_row + 1):
        bottom_left_cell = ws.cell(bottom_row, col)
        bottom_right_cell = ws.cell(bottom_row, top_right)
        if is_blue_header(bottom_left_cell):
            return  bottom_left, bottom_right

        if (has_bottom_border(bottom_left_cell) and
            has_bottom_border(bottom_right_cell)):
            
            bottom_left = (bottom_row, col)
            bottom_right = (bottom_row, top_right)
            return find_last_border(ws, col, max_row, bottom_row, top_right, bottom_left, bottom_right)
    return bottom_left, bottom_right
            


def detect_tables(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    tables = []

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if col != ws.min_column:
                prev_left_cell = ws.cell(row, col - 1)

            # Top-left corner check
            if is_top_left(cell) and is_blue_header(cell) and not is_blue_header(prev_left_cell):
                top_left = (row, col)
                # Find top-right corner
                for right_col in range(col + 1, max_col + 1):
                    right_cell = ws.cell(row, right_col)
                    if has_right_border(right_cell) and is_blue_header(right_cell):

                        next_right_cell = ws.cell(row, right_col + 1)
                        if not is_blue_header(next_right_cell):
                            top_right = (row, right_col)
                            break

                else:
                    continue  # Skip to next cell if no top-right corner found

                # Find bottom-left and bottom-right corners
                for bottom_row in range(row + 1, max_row + 1):
                    bottom_left_cell = ws.cell(bottom_row, col)
                    bottom_right_cell = ws.cell(bottom_row, top_right[1])
                    
                    if (has_bottom_border(bottom_left_cell) and
                        has_bottom_border(bottom_right_cell)):

                        bottom_left = (bottom_row, col)
                        bottom_right = (bottom_row, top_right[1])
                        bottom_left , bottom_right = find_last_border(ws, col, max_row, bottom_row, top_right[1], bottom_left, bottom_right)

                        # Extract the data from the table
                        if bottom_right:
                            table_df = extract_table_data(ws, top_left, bottom_right)
                            tables.append(table_df)
                        
                        break
                    elif is_blue_header(ws.cell(bottom_row + 1, col)):
                        break  # Stop if another header is detected below

    return tables

def convert_datetime_to_string(x):
    if isinstance(x, (datetime, date)):
        return x.isoformat()
    return x

def clean_data_frame(df):
    # Remove rows where all elements are None
    df = df.applymap(convert_datetime_to_string)
    
    df = df.dropna(how='all')
    
    # Remove columns where all elements are None
    df = df.dropna(axis=1, how='all')
    
    # Setting the first row as the header and dropping the first row
    new_header = df.iloc[0] 
    df = df[1:] 
    df.columns = new_header
    
    return df

import json


def dataframe_to_json(df):
    json_data = df.to_json(orient='records', date_format='iso')
    
    metadata = {
        'columns': list(df.columns),
    }
    result = {
        'metadata': metadata,
        'data': json.loads(json_data)
    }
    
    return result


def extract_table_data(ws, top_left, bottom_right):
    """Extract data from a defined range in an Excel worksheet into a DataFrame."""
    data = []
    for row in range(top_left[0], bottom_right[0] + 1):
        row_data = []
        for col in range(top_left[1], bottom_right[1] + 1):
            cell = ws.cell(row, col)
            row_data.append(cell.value)
        data.append(row_data)
        df = pd.DataFrame(data)
    return dataframe_to_json(clean_data_frame(df))



def ask_chatgpt(question, context):

    response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Answer the question based on the context below, and if the question can't be answered based on the context, say \"I don't know\"\n\n"},
                {"role": "user", f"content": f"Context: {context}\n\n---\n\nQuestion: {question}\nAnswer:"}
            ],
            temperature=0,

            top_p=1,
            frequency_penalty=0,
            presence_penalty=0,

        )
    return response.choices[0].message.content

@app.post("/upload/")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="File format not supported.")
    try:
        file_location = f"{file.filename}"
        with open(file_location, "wb") as buffer:
            buffer.write(await file.read())

        tables = detect_tables(file_location)
      
        with open('data.json', 'w') as f:
            json.dump(tables, f, ensure_ascii=False, indent=4)
        return {"message": "File uploaded and processed successfully.", "tables_detected": len(tables)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/ask/")
def ask_question(question: str):
    try:
        with open('data.json', 'r') as file:
            data = json.load(file)
     
        answer = ask_chatgpt(question, data)
        print(answer)
        return {"question": question, "answer": answer}
    except FileNotFoundError as e:
        raise HTTPException(status_code=400,  detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
