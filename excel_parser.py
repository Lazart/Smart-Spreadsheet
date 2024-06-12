import pandas as  pd
from openpyxl import load_workbook

def detect_tables(workbook_path, sheet_name='Sheet1'):
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb.active
    def is_table_start(cell):

        if cell.font.color.type != 'rgb':
            return

        fill_color = cell.fill.start_color.rgb if cell.fill.start_color.type == 'rgb' else cell.fill.start_color.index
        font_color = cell.font.color.rgb if cell.font.color.type == 'rgb' else cell.font.color.index

        is_blue_background = fill_color == 4  
        is_white_font = font_color == 'FFFFFFFF'

        return is_white_font and is_blue_background


    def has_relevant_border(cell, side):
        """ Check if the cell has a relevant border on a specified side ('left' or 'right'). """
        border = cell.border
        if side == 'left':
            return border.left.style is not None
        elif side == 'right':
            return border.right.style is not None
        return False

  
    tables = []
    current_table = []
    table_active = False
    start_col_index = None
    end_col_index = None

    for row in ws.iter_rows():
        row_has_header = any(is_table_start(cell) for cell in row)

        if row_has_header:
            if current_table:  # Finalize the previous table if there is one
                tables.append(current_table)
                current_table = []
            table_active = True
            start_col_index = next((i for i, cell in enumerate(row) if is_table_start(cell)), None)
        
        if table_active:
            end_col_index = next((i for i in range(len(row)-1, -1, -1) if has_relevant_border(row[i], 'right')), None)
        
        if start_col_index is not None and end_col_index is not None:
            # Collect row data between the detected left and right borders
            current_table.append([cell.value for cell in row[start_col_index:end_col_index+1]])

    if current_table:  # Append the last table if not already added
        tables.append(current_table)

    # Convert lists of lists (tables) to DataFrames for easier handling
    table_data = [pd.DataFrame(data) for data in tables]
    return table_data

def clean_column_name(col, default_prefix="Column"):
    """ Clean and standardize column names. """
    if pd.isna(col) or not str(col).strip():
        return default_prefix
    return str(col).strip().replace(" ", "_").replace("/", "_").lower()

def make_columns_unique(df):
    """Ensure DataFrame column names are unique."""
    cols = pd.Series(df.columns)
    seen = {}
    for i, col in enumerate(cols):
        if col in seen:
            seen[col] += 1
            cols[i] = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
    df.columns = cols
    return df

def convert_to_dataframes(tables):
    """ Convert table data to pandas DataFrames and JSON. """
    json_data = []
    for table in tables:
        df = pd.DataFrame(table)
        if not df.empty:
            df.columns = [clean_column_name(col) for col in df.iloc[0]]
            df = df[1:]
            df = make_columns_unique(df)
            json_output = df.to_json(orient='records', date_format='iso')
            json_data.append(json_output)
    return json_data

# Load and process the Excel file.
file_path = './tests/example_2.xlsx'
